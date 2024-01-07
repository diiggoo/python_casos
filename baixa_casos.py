import time
import winreg
import re
import platform
import os
import zipfile
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
from datetime import datetime

def get_chrome_version():
    try:
        # Caminho do registro onde as informações do Google Chrome estão armazenadas
        reg_path = r"SOFTWARE\Google\Chrome\BLBeacon"

        # Abre a chave do registro
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path)

        # Obtém o valor da chave "version"
        chrome_version, _ = winreg.QueryValueEx(key, "version")

        # Fecha a chave do registro
        winreg.CloseKey(key)

        # Usa expressão regular para extrair apenas os dígitos antes do primeiro ponto
        match = re.match(r'(\d+)', chrome_version)
        chrome_version_digits = match.group(1) if match else "Não foi possível extrair a versão."

        return chrome_version_digits

    except Exception as e:
        return f"Erro ao obter a versão do Google Chrome: {e}"

def get_os_architecture():
    try:
        os_architecture = platform.architecture()[0]
        return 'win64' if os_architecture == '64bit' else 'win32'

    except Exception as e:
        return f"Erro ao obter a arquitetura do sistema operacional: {e}"

def download_and_install_chromedriver(version_chrome, version_so):
    chromedriver_url = 'https://googlechromelabs.github.io/chrome-for-testing/'

    try:
        response = requests.get(chromedriver_url, verify=False)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Erro na solicitação HTTP: {e}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')
    pattern = re.compile(fr'https://edgedl\.me\.gvt1\.com/edgedl/chrome/chrome-for-testing/{version_chrome}.*chromedriver-{version_so}.zip')

    download_folder = 'drive'  # Nome da pasta onde o arquivo ZIP será baixado
    create_folder = True  # Flag para indicar se a pasta 'drive' precisa ser criada

    for section in soup.find_all('section', id='stable'):
        for link in section.find_all('code'):
            result = str(link)
            match = pattern.search(result)
            if match:
                chromedriver_link = match.group(0)
                download_path = os.path.join(download_folder, f'chromedriver_{version_chrome}_{version_so}.zip')

                try:
                    # Cria a pasta 'drive' se não existir
                    if create_folder:
                        os.makedirs(download_folder, exist_ok=True)
                        create_folder = False

                    # Baixa o arquivo zip do chromedriver
                    response = requests.get(chromedriver_link, stream=True)
                    response.raise_for_status()
                    with open(download_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            f.write(chunk)
                    
                    print(f"Chromedriver baixado com sucesso em: {download_path}")

                    # Descompacta o arquivo ZIP diretamente na pasta 'drive'
                    with zipfile.ZipFile(download_path, 'r') as zip_ref:
                        zip_ref.extractall(download_folder)

                    print(f"Chromedriver descompactado com sucesso em: {download_folder}")

                except requests.exceptions.RequestException as e:
                    print(f"Erro ao baixar o Chromedriver: {e}")
                except zipfile.BadZipFile:
                    print(f"Erro: O arquivo baixado não é um arquivo ZIP válido.")
                except Exception as e:
                    print(f"Erro ao processar o Chromedriver: {e}")

                return

    print(f"Chromedriver não encontrado para a versão {version_chrome} e SO {version_so}")

#Caminho driver
chrome_driver_path = r'drive\chromedriver-win64\chromedriver.exe'

# Configuração do ChromeDriver
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--start-maximized')  # Inicia o navegador maximizado (opcional)

if __name__ == "__main__":

    current_datetime = datetime.now().strftime("%d/%m/%Y")
    try: 
        # Inicializa o navegador Chrome
        service = Service(executable_path=chrome_driver_path)
        options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(service=service, options=options)
        driver.maximize_window()

        # Exemplo: Abre o site do Google
        driver.get("")
        time.sleep(0.5)
        username = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="email"]')))
        username.send_keys('')
        time.sleep(0.5)

        element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="proximo"]')))
        element.click()
        
        senha = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="password"]')))
        senha.send_keys('')

        element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="btnEnviarForm"]')))
        element.click()

        time.sleep(2)

        element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="inter"]/div/div/div/a[2]')))
        element.click()



        # Feche o navegador
        #driver.quit()

    except Exception as e:
        print(f"Driver não encontrado: ", e)
        chrome_version = get_chrome_version()
        os_architecture = get_os_architecture()
        download_and_install_chromedriver(chrome_version, os_architecture)
        print(f"Versão do Google Chrome instalado: {chrome_version}")
        print(f"Arquitetura do sistema operacional: {os_architecture}")

    #Abre o arquivo Excel
    excel_file_path = 'PROJETO BAIXA.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)

    #Selecione a planilha desejada (por nome)
    sheet_name = 'Planilha2'
    sheet = workbook[sheet_name]

    #Obtenha as dimensões da planilha (número de linhas e colunas)
    max_row = sheet.max_row
    max_column = sheet.max_column

    #Imprime o conteúdo da planilha
    for row in sheet.iter_rows(min_row=1, max_row=max_row, values_only=True):
        if row[89] == 'FINALIZADO' or row[89] == 'ARQUIVADO':
            numero_processo = row[19]
            valor_processo = round(row[85],2)
            if valor_processo == 0 or valor_processo >= 999999:
                valor_processo = 0.01

            valor_processo_formatado = f"{valor_processo:.2f}".replace('.', ',')
            print(f'Status do processo {row[89]} Numero do processo {numero_processo} e valor do processo {valor_processo_formatado}')

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="clean_filtro1"]/button')))
            element.click()

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="filtro1"]')))
            element.send_keys(numero_processo)

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="icon_filtro1"]')))
            element.click()

            time.sleep(3)
            
            try:
                span_edit = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//*[@id="tbControle"]/tbody/tr/td[1]/span[1]')))
                span_edit.click()
                time.sleep(3)


            except:
                element = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.XPATH, '//*[@id="clean_filtro1"]/button')))
                element.click()
                continue           

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="saida"]')))
            element.click()

            element.send_keys(current_datetime)

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="status"]')))
            element.click()
            drop = Select(element)
            drop.select_by_visible_text('FINALIZADO')

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="empreitada"]')))
            element.click()

            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="vrNossoCalculo"]')))
            element.send_keys(valor_processo_formatado)            
            
            time.sleep(3)
            element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="btnEnviarForm"]')))
            element.click()
